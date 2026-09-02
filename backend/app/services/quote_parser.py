from __future__ import annotations

import csv
import io
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

import pdfplumber
from openpyxl import load_workbook
from PIL import Image, ImageEnhance, ImageOps, UnidentifiedImageError
from pypdf import PdfReader
from pypdf.errors import PdfReadError


PARSER_VERSION = "beta-1"
MAX_OCR_PAGES = 20
ALIASES = {
    "name": {"项目", "项目名称", "名称", "施工项目", "工程项目", "装饰装修内容及材料"},
    "area": {"区域", "空间", "房间", "施工区域"},
    "category": {"类别", "分类", "施工类别", "专业"},
    "quantity": {"数量", "工程量"},
    "unit": {"单位", "计量单位"},
    "unit_price": {"单价", "综合单价"},
    "total": {"合价", "金额", "总价", "小计", "合计金额", "综合合价"},
    "material": {"材料", "材料品牌型号", "品牌型号", "品牌规格", "规格型号品牌等级"},
    "craft": {"工艺", "工艺说明", "施工工艺", "项目特征", "做法说明"},
}
CATEGORY_KEYWORDS = {
    "水电": ("电线", "电路", "插座", "开关", "给水", "排水", "水管", "强电", "弱电"),
    "泥瓦": ("铺贴", "瓷砖", "地砖", "墙砖", "砌墙", "找平", "水泥", "砂浆"),
    "油漆": ("乳胶漆", "腻子", "涂料", "油漆", "墙面基层"),
    "防水": ("防水", "闭水"),
    "木作": ("吊顶", "石膏板", "柜", "木饰面", "龙骨"),
    "拆除": ("拆除", "铲除", "拆墙"),
    "安装": ("安装", "灯具", "洁具", "五金"),
}
UNIT_PATTERN = r"㎡|m²|m2|mi|ni|m|米|延米|个|位|项|套|樘|组|块|台"
LINE_ITEM_PATTERN = re.compile(
    rf"^(?P<prefix>.*?)\s+(?P<quantity>\d+(?:\.\d+)?)\s*(?P<unit>{UNIT_PATTERN})\s+"
    r"(?P<unit_price>[\d,]+(?:\.\d+)?)\s+(?P<total>[\d,]+(?:\.\d+)?)\s*$",
    re.IGNORECASE,
)


@dataclass
class ParseResult:
    items: list[dict]
    input_type: str
    parse_method: str
    page_count: int = 1
    source_total_cents: int | None = None
    warnings: list[str] = field(default_factory=list)
    parser_version: str = PARSER_VERSION


def _money(value) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).replace("¥", "").replace("￥", "").replace(",", "").strip()
    try:
        return round(float(text) * 100)
    except ValueError:
        return None


def _decimal(value) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).replace(",", "").strip()
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return Decimal(match.group())
    except InvalidOperation:
        return None


def _header_key(value) -> str | None:
    label = re.sub(r"\s+", "", str(value or "")).strip()
    if not label:
        return None
    for key, aliases in ALIASES.items():
        if label in aliases or any(alias in label for alias in aliases if len(alias) >= 2):
            return key
    return None


def _infer_category(name: str, category: str | None = None) -> str:
    if category and category.strip() and category.strip() != "其他":
        return category.strip()
    for candidate, keywords in CATEGORY_KEYWORDS.items():
        if any(keyword in name for keyword in keywords):
            return candidate
    return "其他"


def standardize_item_name(name: str) -> str:
    normalized = re.sub(r"[（）()【】\[\]·、，,：:\s]+", "", name).lower()
    replacements = {
        "墙顶面乳胶漆": "乳胶漆涂刷",
        "墙面乳胶漆": "乳胶漆涂刷",
        "电路点位改造": "强电点位",
        "装修垃圾清运费": "垃圾清运",
        "轻钢龙骨石膏板吊顶": "石膏板吊顶",
    }
    return replacements.get(normalized, name.strip())


def _normalize(rows: Iterable[list], sheet: str, confidence: int = 100) -> list[dict]:
    values = [list(row) for row in rows]
    if not values:
        return []
    header_index = 0
    mapping: dict[str, int] = {}
    for row_index, row in enumerate(values[:25]):
        trial: dict[str, int] = {}
        for column, cell in enumerate(row):
            key = _header_key(cell)
            if key and key not in trial:
                trial[key] = column
        if "name" in trial and "total" in trial:
            header_index, mapping = row_index, trial
            break
    if not mapping:
        raise ValueError("未找到包含“项目名称”和“合价/金额”的表头")
    result = []
    for row_number, row in enumerate(values[header_index + 1 :], header_index + 2):
        def value(key):
            index = mapping.get(key)
            return row[index] if index is not None and index < len(row) else None

        name = str(value("name") or "").strip()
        total = _money(value("total"))
        if not name or total is None or any(word in name for word in ("合计", "总计")):
            continue
        standard_name = standardize_item_name(name)
        category = _infer_category(standard_name, str(value("category") or ""))
        result.append(
            {
                "original_name": name,
                "standard_name": standard_name,
                "area": str(value("area") or "").strip() or None,
                "category": category,
                "quantity_text": str(value("quantity") or "").strip() or None,
                "unit": str(value("unit") or "").strip() or None,
                "unit_price_cents": _money(value("unit_price")),
                "total_cents": total,
                "material_info": str(value("material") or "").strip() or None,
                "craft_notes": str(value("craft") or "").strip() or None,
                "source_location": f"{sheet}!第 {row_number} 行",
                "source_excerpt": " | ".join(str(cell) for cell in row if cell not in (None, ""))[:800],
                "confidence": confidence,
                "field_confidences": {
                    "standard_name": confidence,
                    "quantity": confidence if value("quantity") not in (None, "") else 0,
                    "unit_price_cents": confidence if value("unit_price") not in (None, "") else 0,
                    "total_cents": confidence,
                },
            }
        )
    if not result:
        raise ValueError("表格中没有可导入的有效报价条目")
    return result


def _parse_spreadsheet(content: bytes, suffix: str) -> ParseResult:
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        items = _normalize(csv.reader(io.StringIO(text)), "CSV")
        return ParseResult(items=items, input_type="csv", parse_method="deterministic_table")
    book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    items: list[dict] = []
    for sheet in book.worksheets:
        try:
            items.extend(_normalize(sheet.iter_rows(values_only=True), sheet.title))
        except ValueError:
            continue
    if not items:
        raise ValueError("工作簿中没有找到可识别的报价表")
    return ParseResult(items=items, input_type="xlsx", parse_method="deterministic_table")


def _clean_item_prefix(prefix: str) -> str:
    return re.sub(r"^\s*(?:\d+(?:\.\d+)*|[A-Za-z]\d+)\s*", "", prefix).strip(" |-—")


def parse_text_rows(text: str, page_number: int = 1, confidence: int = 88, method: str = "PDF 文本") -> list[dict]:
    items: list[dict] = []
    for line_number, raw in enumerate(text.splitlines(), 1):
        line = re.sub(r"\s+", " ", raw).strip()
        if not line or any(header in line for header in ("项目名称 数量", "序号 项目", "单价 合价")):
            continue
        match = LINE_ITEM_PATTERN.match(line)
        if not match:
            continue
        name = _clean_item_prefix(match.group("prefix"))
        if len(name) < 2 or any(word in name for word in ("合计", "总计", "小计")):
            continue
        total = _money(match.group("total"))
        if total is None:
            continue
        standard_name = standardize_item_name(name)
        unit = match.group("unit")
        if unit.lower() in {"m²", "m2", "mi", "ni"}:
            unit = "㎡"
        items.append(
            {
                "original_name": name,
                "standard_name": standard_name,
                "area": None,
                "category": _infer_category(standard_name),
                "quantity_text": match.group("quantity"),
                "unit": unit,
                "unit_price_cents": _money(match.group("unit_price")),
                "total_cents": total,
                "material_info": None,
                "craft_notes": None,
                "source_location": f"{method}第 {page_number} 页·行 {line_number}",
                "source_excerpt": line[:800],
                "confidence": confidence,
                "field_confidences": {"standard_name": confidence, "quantity": confidence, "unit_price_cents": confidence, "total_cents": confidence},
            }
        )
    return items


def _extract_source_total(text: str) -> int | None:
    matches = re.findall(r"(?:报价总计|工程造价|总计|合计)\s*[：:]?\s*[¥￥]?\s*([\d,]+(?:\.\d{1,2})?)", text)
    return _money(matches[-1]) if matches else None


def _parse_pdf_tables(content: bytes) -> list[dict]:
    items: list[dict] = []
    with pdfplumber.open(io.BytesIO(content)) as document:
        for page_number, page in enumerate(document.pages, 1):
            for table_index, table in enumerate(page.extract_tables() or [], 1):
                try:
                    items.extend(_normalize(table, f"PDF 第 {page_number} 页·表 {table_index}", confidence=92))
                except ValueError:
                    continue
    return items


def _ocr_language() -> str:
    completed = subprocess.run(["tesseract", "--list-langs"], capture_output=True, text=True, check=True)
    available = set(completed.stdout.split())
    if "chi_sim" in available:
        return "chi_sim+eng" if "eng" in available else "chi_sim"
    if "eng" in available:
        return "eng"
    raise ValueError("OCR 未安装可用语言包，请安装 tesseract-lang")


def _prepare_image(source: Path, target: Path) -> None:
    try:
        with Image.open(source) as image:
            normalized = ImageEnhance.Contrast(ImageOps.exif_transpose(image).convert("L")).enhance(1.35)
            if normalized.width < 1600:
                ratio = 1600 / normalized.width
                normalized = normalized.resize((1600, round(normalized.height * ratio)))
            normalized.save(target, format="PNG", optimize=True)
    except (UnidentifiedImageError, OSError) as exc:
        raise ValueError("图片无法读取；HEIC 请先转为 JPG/PNG 后重试") from exc


def _ocr_image(path: Path) -> tuple[str, int]:
    if not shutil.which("tesseract"):
        raise ValueError("当前环境未安装 OCR 引擎 Tesseract")
    completed = subprocess.run(
        ["tesseract", str(path), "stdout", "-l", _ocr_language(), "--psm", "3", "tsv"],
        capture_output=True,
        text=True,
        check=False,
        timeout=45,
    )
    if completed.returncode != 0:
        raise ValueError("OCR 处理失败，请改用更清晰图片或模板导入")
    grouped: dict[tuple[str, str, str, str], list[tuple[str, float]]] = {}
    confidences: list[float] = []
    for row in csv.DictReader(io.StringIO(completed.stdout), delimiter="\t"):
        token = (row.get("text") or "").strip()
        if not token:
            continue
        try:
            token_confidence = float(row.get("conf") or -1)
        except ValueError:
            token_confidence = -1
        key = tuple(row.get(name, "") for name in ("block_num", "par_num", "line_num", "page_num"))
        grouped.setdefault(key, []).append((token, token_confidence))
        if token_confidence >= 0:
            confidences.append(token_confidence)
    text = "\n".join(" ".join(token for token, _ in tokens) for tokens in grouped.values())
    confidence = round(sum(confidences) / len(confidences)) if confidences else 45
    return text, max(1, min(99, confidence))


def _parse_image_or_scanned_pdf(content: bytes, suffix: str, page_count: int) -> ParseResult:
    warnings: list[str] = []
    items: list[dict] = []
    with tempfile.TemporaryDirectory(prefix="quote-ocr-") as directory:
        temp_dir = Path(directory)
        if suffix == ".pdf":
            if not shutil.which("pdftoppm"):
                raise ValueError("当前环境未安装 PDF 渲染组件 Poppler")
            source_pdf = temp_dir / "source.pdf"
            source_pdf.write_bytes(content)
            limit = min(page_count, MAX_OCR_PAGES)
            completed = subprocess.run(
                ["pdftoppm", "-f", "1", "-l", str(limit), "-r", "200", "-jpeg", str(source_pdf), str(temp_dir / "page")],
                capture_output=True,
                text=True,
                check=False,
                timeout=90,
            )
            if completed.returncode != 0:
                raise ValueError("扫描 PDF 无法转换为图片")
            sources = sorted(temp_dir.glob("page-*.jpg"))
            if page_count > MAX_OCR_PAGES:
                warnings.append(f"扫描 PDF 共 {page_count} 页，本次仅识别前 {MAX_OCR_PAGES} 页")
        else:
            source_image = temp_dir / f"source{suffix}"
            source_image.write_bytes(content)
            sources = [source_image]
        for page_number, source in enumerate(sources, 1):
            prepared = temp_dir / f"prepared-{page_number}.png"
            _prepare_image(source, prepared)
            text, confidence = _ocr_image(prepared)
            items.extend(parse_text_rows(text, page_number, min(confidence, 72), method="OCR "))
    if not items:
        raise ValueError("未能从扫描件中识别出完整报价行；请使用模板导入或手工录入")
    return ParseResult(items=items, input_type="scanned_pdf" if suffix == ".pdf" else "image", parse_method="ocr_tesseract", page_count=page_count, warnings=warnings)


def _normalize_ai_items(raw_items: list[dict]) -> tuple[list[dict], list[str]]:
    items: list[dict] = []
    warnings: list[str] = []
    arithmetic_mismatches = 0
    for raw in raw_items:
        name = str(raw.get("original_name") or "").strip()
        total = _money(raw.get("total"))
        if not name or total is None or any(word in name for word in ("合计", "总计", "小计")):
            continue
        quantity_text = str(raw.get("quantity") or "").strip() or None
        unit_price = _money(raw.get("unit_price"))
        quantity = _decimal(quantity_text)
        model_confidence = raw.get("confidence")
        confidence = 90 if quantity_text and raw.get("unit") and unit_price is not None else 80
        if isinstance(model_confidence, int):
            confidence = min(confidence, max(35, model_confidence))
        arithmetic_ok = True
        if quantity is not None and unit_price is not None:
            calculated = round(quantity * unit_price)
            tolerance = max(2, round(abs(total) * 0.005))
            arithmetic_ok = abs(calculated - total) <= tolerance
            if not arithmetic_ok:
                confidence = min(confidence, 65)
                arithmetic_mismatches += 1
        standard_name = standardize_item_name(name)
        source_excerpt = str(raw.get("source_excerpt") or "").strip()
        if not source_excerpt:
            source_excerpt = " | ".join(
                str(value) for value in (name, quantity_text, raw.get("unit"), raw.get("unit_price"), raw.get("total")) if value not in (None, "")
            )
        page_number = int(raw.get("page_number") or 1)
        item_number = int(raw.get("item_number") or len(items) + 1)
        items.append(
            {
                "original_name": name,
                "standard_name": standard_name,
                "area": str(raw.get("area") or "").strip() or None,
                "category": _infer_category(standard_name, str(raw.get("category") or "")),
                "quantity_text": quantity_text,
                "unit": str(raw.get("unit") or "").strip() or None,
                "unit_price_cents": unit_price,
                "total_cents": total,
                "material_info": str(raw.get("material_info") or "").strip() or None,
                "craft_notes": str(raw.get("craft_notes") or "").strip() or None,
                "source_location": f"AI 视觉第 {page_number} 页·条目 {item_number}",
                "source_excerpt": source_excerpt[:800],
                "confidence": confidence,
                "field_confidences": {
                    "standard_name": confidence,
                    "quantity": confidence if quantity_text else 0,
                    "unit_price_cents": confidence if unit_price is not None else 0,
                    "total_cents": confidence if arithmetic_ok else min(confidence, 65),
                },
            }
        )
    if arithmetic_mismatches:
        warnings.append(f"{arithmetic_mismatches} 个 AI 条目的数量×单价与合价不一致，已标记为低置信度")
    if not items:
        raise ValueError("AI 模型没有返回可入库的完整报价条目")
    return items, warnings


def _parse_with_ai(content: bytes, suffix: str, page_count: int) -> ParseResult:
    from app.core.config import get_settings
    from app.services.quote_ai import QwenQuoteParser

    settings = get_settings()
    with QwenQuoteParser(
        api_key=settings.dashscope_api_key,
        base_url=settings.dashscope_base_url,
        primary_model=settings.qwen_primary_model,
        fallback_model=settings.qwen_fallback_model,
        ocr_model=settings.qwen_ocr_model,
        timeout_seconds=settings.ai_request_timeout_seconds,
        max_retries=settings.ai_max_retries,
        max_pages=settings.ai_max_pages,
    ) as parser:
        result = parser.parse_document(content, suffix, page_count)
    items, validation_warnings = _normalize_ai_items(result.items)
    return ParseResult(
        items=items,
        input_type="scanned_pdf" if suffix == ".pdf" else "image",
        parse_method=result.parse_method,
        page_count=page_count,
        source_total_cents=_money(result.source_total),
        warnings=[*result.warnings, *validation_warnings],
        parser_version=result.parser_version,
    )


def _parse_visual_with_fallback(content: bytes, suffix: str, page_count: int) -> ParseResult:
    from app.core.config import get_settings
    from app.services.quote_ai import QuoteAIError

    settings = get_settings()
    if not settings.ai_ready:
        return _parse_image_or_scanned_pdf(content, suffix, page_count)
    try:
        return _parse_with_ai(content, suffix, page_count)
    except (QuoteAIError, ValueError) as ai_error:
        try:
            local = _parse_image_or_scanned_pdf(content, suffix, page_count)
        except ValueError as local_error:
            raise ValueError(f"{ai_error}；本地 OCR 回退也失败：{local_error}") from local_error
        local.parse_method = "ocr_tesseract_fallback"
        local.warnings.insert(0, f"AI 服务未完成解析，已自动使用本地 OCR：{ai_error}")
        return local


def _parse_pdf(content: bytes) -> ParseResult:
    reader = PdfReader(io.BytesIO(content))
    if reader.is_encrypted:
        try:
            reader.decrypt("")
        except Exception as exc:
            raise ValueError("PDF 已加密，无法解析") from exc
    page_count = len(reader.pages)
    page_texts = [(page.extract_text() or "") for page in reader.pages]
    combined = "\n".join(page_texts)
    if len(re.sub(r"\s+", "", combined)) >= 40:
        table_items = _parse_pdf_tables(content)
        if table_items:
            return ParseResult(items=table_items, input_type="text_pdf", parse_method="pdf_table", page_count=page_count, source_total_cents=_extract_source_total(combined))
        line_items: list[dict] = []
        for page_number, text in enumerate(page_texts, 1):
            line_items.extend(parse_text_rows(text, page_number))
        if line_items:
            return ParseResult(items=line_items, input_type="text_pdf", parse_method="pdf_text", page_count=page_count, source_total_cents=_extract_source_total(combined))
    return _parse_visual_with_fallback(content, ".pdf", page_count)


def parse_quote_document(content: bytes, suffix: str) -> ParseResult:
    suffix = suffix.lower()
    try:
        if suffix in {".csv", ".xlsx"}:
            return _parse_spreadsheet(content, suffix)
        if suffix == ".pdf":
            return _parse_pdf(content)
        if suffix in {".jpg", ".jpeg", ".png", ".heic"}:
            return _parse_visual_with_fallback(content, suffix, 1)
        raise ValueError("支持 XLSX、CSV、文本 PDF、扫描 PDF、JPG、JPEG、PNG 和 HEIC")
    except ValueError:
        raise
    except subprocess.TimeoutExpired as exc:
        raise ValueError("文件识别超时，请压缩文件、拆分页数或改用模板导入") from exc
    except (OSError, EOFError, PdfReadError) as exc:
        raise ValueError("文件无法读取或内容已损坏") from exc


def parse_quote(content: bytes, suffix: str) -> list[dict]:
    """Backward-compatible item-only parser used by existing callers and tests."""
    return parse_quote_document(content, suffix).items
