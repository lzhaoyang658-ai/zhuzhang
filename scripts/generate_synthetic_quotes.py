from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


OUTPUT_DIR = Path("backend/data/sample_quotes/synthetic")
HEADERS = ["序号", "项目名称", "区域", "施工类别", "数量", "单位", "单价", "合价", "材料品牌型号", "工艺说明"]

BASE_ITEMS = [
    ("墙面基层处理", "客餐厅", "油漆", 86, "㎡", 38, "立邦净味腻子", "基层清理、两遍腻子并打磨"),
    ("乳胶漆涂刷", "客餐厅", "油漆", 86, "㎡", 52, "立邦金装五合一", "一底两面，成品保护"),
    ("地砖铺贴", "客餐厅", "泥瓦", 42, "㎡", 128, "东鹏 800×800", "含水泥砂浆辅料，不含主材"),
    ("防水施工", "卫生间", "防水", 18, "㎡", 95, "东方雨虹柔性防水", "墙地面两遍，闭水 48 小时"),
    ("给水管改造", "全屋", "水电", 68, "m", 88, "伟星 PPR 25", "热熔连接，含开槽及封槽"),
    ("强电点位", "全屋", "水电", 52, "位", 145, "远东电缆+施耐德底盒", "2.5/4 平方线分路敷设"),
    ("石膏板吊顶", "客餐厅", "木作", 28, "㎡", 165, "可耐福石膏板", "轻钢龙骨，转角整板处理"),
    ("垃圾清运", "全屋", "其他", 1, "项", 2600, "", "袋装清运至物业指定点"),
]


def vendor_items(price_factor: float, variants: dict[str, tuple[str, float]] | None = None):
    variants = variants or {}
    result = []
    for index, (name, area, category, quantity, unit, unit_price, material, craft) in enumerate(BASE_ITEMS, 1):
        final_name, final_factor = variants.get(name, (name, 1.0))
        price = round(unit_price * price_factor * final_factor, 2)
        total = round(quantity * price, 2)
        result.append({
            "序号": index,
            "项目名称": final_name,
            "区域": area,
            "施工类别": category,
            "数量": quantity,
            "单位": unit,
            "单价": price,
            "合价": total,
            "材料品牌型号": material,
            "工艺说明": craft,
        })
    return result


def write_xlsx(path: Path, title: str, items: list[dict]):
    book = Workbook()
    sheet = book.active
    sheet.title = "报价明细"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    sheet.cell(1, 1, title).font = Font(size=16, bold=True)
    sheet.cell(1, 1).alignment = Alignment(horizontal="center")
    for column, header in enumerate(HEADERS, 1):
        cell = sheet.cell(3, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="365D52")
        cell.alignment = Alignment(horizontal="center")
    for row_index, item in enumerate(items, 4):
        for column, header in enumerate(HEADERS, 1):
            sheet.cell(row_index, column, item[header])
    total_row = 4 + len(items)
    sheet.cell(total_row, 2, "报价总计")
    sheet.cell(total_row, 8, f"=SUM(H4:H{total_row - 1})")
    sheet.cell(total_row, 2).font = sheet.cell(total_row, 8).font = Font(bold=True)
    widths = [8, 22, 14, 14, 10, 10, 12, 14, 24, 38]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A4"
    book.save(path)


def write_csv(path: Path, items: list[dict]):
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(items)


def write_scan_image(path: Path, title: str, items: list[dict]):
    canvas = Image.new("RGB", (2400, 1540), "#f5f3ed")
    draw = ImageDraw.Draw(canvas)
    font_candidates = [
        Path("/System/Library/Fonts/Hiragino Sans GB.ttc"),
        Path("/System/Library/Fonts/STHeiti Medium.ttc"),
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
    ]
    font_path = next((candidate for candidate in font_candidates if candidate.exists()), None)
    if font_path is None:
        raise RuntimeError("未找到可用于生成中文扫描样本的字体")
    title_font = ImageFont.truetype(str(font_path), 52)
    header_font = ImageFont.truetype(str(font_path), 30)
    row_font = ImageFont.truetype(str(font_path), 29)
    draw.text((100, 70), title, fill="#1e2925", font=title_font)
    draw.text((100, 160), "项目名称    空间    类别    数量 单位    单价    合价", fill="#274f43", font=header_font)
    y = 235
    for item in items:
        line = f'{item["项目名称"]}    {item["区域"]}    {item["施工类别"]}    {item["数量"]} {item["单位"]}    {item["单价"]}    {item["合价"]}'
        draw.text((100, y), line, fill="#252b28", font=row_font)
        draw.line((100, y + 48, 2270, y + 48), fill="#d4d0c7", width=2)
        y += 120
    total = sum(item["合价"] for item in items)
    draw.text((100, y + 20), f"报价总计    {total:.2f}", fill="#1d473a", font=header_font)
    canvas.save(path, quality=94, subsampling=0)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    vendors = {
        "vendor-a": vendor_items(1.0),
        "vendor-b": vendor_items(0.96, {"乳胶漆涂刷": ("墙顶面乳胶漆", 1.05), "强电点位": ("电路点位改造", 1.08)}),
        "vendor-c": vendor_items(1.04, {"垃圾清运": ("装修垃圾清运费", 0.9), "石膏板吊顶": ("轻钢龙骨石膏板吊顶", 1.02)}),
    }
    write_xlsx(OUTPUT_DIR / "vendor-a.xlsx", "虚构样本 A - 青禾装饰报价", vendors["vendor-a"])
    write_xlsx(OUTPUT_DIR / "vendor-b.xlsx", "虚构样本 B - 木川空间报价", vendors["vendor-b"])
    write_csv(OUTPUT_DIR / "vendor-c.csv", vendors["vendor-c"])
    write_scan_image(OUTPUT_DIR / "vendor-a-scan.jpg", "虚构样本 A 装修工程报价单", vendors["vendor-a"])
    truth = {
        key: {
            "item_count": len(items),
            "total_cents": round(sum(item["合价"] for item in items) * 100),
            "items": items,
        }
        for key, items in vendors.items()
    }
    (OUTPUT_DIR / "ground-truth.json").write_text(json.dumps(truth, ensure_ascii=False, indent=2), encoding="utf-8")
    for key, value in truth.items():
        print(f"generated {key}: {value['item_count']} items, {value['total_cents']} cents")


if __name__ == "__main__":
    main()
